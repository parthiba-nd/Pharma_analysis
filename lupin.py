import pyodbc
import pandas as pd
from google.cloud import firestore, storage
from datetime import datetime, date, timedelta
import json
import re
import warnings
warnings.filterwarnings('ignore')

# ------------------------------------------------------------
# 1️⃣ LOAD EMPLOYEE MASTER FROM FIREBASE
# ------------------------------------------------------------
st = storage.Client("neodocs-8d6cd")
bucket = st.bucket("neodocs-8d6cd-utils")

blob = bucket.blob("org_access_codes/lupin-hb.json")
json_data = blob.download_as_text()
employee_dict = json.loads(json_data)

# Convert to DataFrame
emp_df = pd.DataFrame(employee_dict).transpose().reset_index()
emp_df.rename(columns={"index": "empId", "mr_region": "hq"}, inplace=True)

# Remove Training rows
emp_df = emp_df[emp_df['mr_name'] != "Training"]

# Convert region_list to comma-separated text
if "region_list" in emp_df.columns:
    emp_df["region_list"] = emp_df["region_list"].apply(
        lambda x: ", ".join(x) if isinstance(x, list) else x
    )

# Ensure empId is string
emp_df["empId"] = emp_df["empId"].astype(str)

print("Employee Master Loaded:", emp_df.shape)

# ------------------------------------------------------------
# 2️⃣ SQL SERVER — LOAD DOCTOR METRICS
# ------------------------------------------------------------
server = 'neodocs-sql-server.database.windows.net'
database = 'neodocs-sql-db'
username = 'ndDashboard'
password = 'NeoDocs@2025'
driver = '{ODBC Driver 18 for SQL Server}'

conn_str = f"""
DRIVER={driver};
SERVER={server};
DATABASE={database};
UID={username};
PWD={password};
Encrypt=yes;
TrustServerCertificate=no;
Connection Timeout=30;
"""

conn = pyodbc.connect(conn_str)
print("Connected to SQL Server")

# ------------------------------------------------------------
# DATE RANGE — CURRENT MONTH
# ------------------------------------------------------------
today = date.today()

# First day of current month
first_day_this_month = today.replace(day=1)

# Start = first day of current month
START_DATE = first_day_this_month.strftime('%Y-%m-%d')

# End = today (you can change to end of month if needed)
END_DATE = today.strftime('%Y-%m-%d')

print("SQL Range (Current Month):", START_DATE, "to", END_DATE)


# ------------------------------------------------------------
# CLEAN SQL QUERY (hemaday + lupin_hb HB)
# ------------------------------------------------------------
query = """
WITH test_summary AS (
    SELECT 
        u.empId,
        u.docId AS [Doctor ID],
        u.drName AS Doctor,
        u.oId,
        COUNT(DISTINCT u.testId) AS [Total Tests],
        COUNT(DISTINCT CONVERT(date, u.campDate)) AS [Total Camps]
    FROM dbo.user_tests u
    LEFT JOIN dbo.aId a ON u.aId = a.aId
    WHERE 
        a.aId = '02c6bc8e-9395-4cff-80cc-af0df4f951af'
        AND u.statusCode = 200
        AND u.isDeleted = 0
        AND u.campDate BETWEEN ? AND ?
    GROUP BY u.empId, u.docId, u.drName, u.oId
),

rx_summary AS (
    SELECT 
        r.aId,
        r.oId,
        r.campDate,
        SUM(TRY_CAST(LEFT(JSON_VALUE(r.prescriptions, '$.lupiheme'), 
            CHARINDEX('|', JSON_VALUE(r.prescriptions, '$.lupiheme')) - 1) AS INT)) AS [Total Rx],
        SUM(TRY_CAST(LTRIM(RIGHT(JSON_VALUE(r.prescriptions, '$.lupiheme'), 
            LEN(JSON_VALUE(r.prescriptions, '$.lupiheme')) - 
            CHARINDEX('|', JSON_VALUE(r.prescriptions, '$.lupiheme')))) AS INT)) AS [Total Strips]
    FROM dbo.rx r
    LEFT JOIN dbo.aId a ON r.aId = a.aId
    WHERE 
        a.aId = '02c6bc8e-9395-4cff-80cc-af0df4f951af'
        AND r.isDeleted = 0
        AND r.campDate BETWEEN ? AND ?
    GROUP BY r.aId, r.oId, r.campDate
)

SELECT 
    t.empId AS [empId],
    t.Doctor,
    t.[Doctor ID],
    t.[Total Camps] AS [Doc Total Camps],
    ISNULL(SUM(rx.[Total Rx]), 0) AS [Total Rx],
    ISNULL(SUM(rx.[Total Strips]), 0) AS [Total Strips],
    t.[Total Tests]
FROM test_summary t
LEFT JOIN rx_summary rx
    ON t.oId = rx.oId
GROUP BY 
    t.empId, t.Doctor, t.[Doctor ID], t.[Total Camps], t.[Total Tests]
ORDER BY t.empId;
"""

params = [START_DATE, END_DATE, START_DATE, END_DATE]
sql_df = pd.read_sql(query, conn, params=params)

print("Doctor Metrics Loaded:", sql_df.shape)

# ------------------------------------------------------------
# 3️⃣ ASSIGN DOCTOR INDEX PER empId
# ------------------------------------------------------------
sql_df['Doctor Index'] = sql_df.groupby(['empId']).cumcount() + 1

# ------------------------------------------------------------
# 4️⃣ PIVOT DOCTOR METRICS — Wide format
# ------------------------------------------------------------
pivoted = sql_df.pivot_table(
    index=['empId'],
    columns='Doctor Index',
    values=['Doctor', 'Doctor ID', 'Doc Total Camps', 'Total Rx', 'Total Strips', 'Total Tests'],
    aggfunc='first'
)

# Flatten columns
pivoted.columns = [f"{col[0]} {int(col[1])}" for col in pivoted.columns]
pivoted = pivoted.reset_index()

# ------------------------------------------------------------
# ADD TOTAL CAMPS COLUMN (SUM OF ALL DOC TOTAL CAMPS)
# ------------------------------------------------------------

# Identify all Doc Total Camps columns
doc_camp_cols = [col for col in pivoted.columns if col.startswith("Doc Total Camps")]

# Create Total Camps column
pivoted["Total Camps"] = pivoted[doc_camp_cols].sum(axis=1)


# ------------------------------------------------------------
# REORDER COLUMNS WITH TOTAL CAMPS FIRST, THEN DOCTORS
# ------------------------------------------------------------

# Identify all doctor indices
doctor_indices = sorted(
    list({
        int(col.split()[-1]) 
        for col in pivoted.columns 
        if col not in ["empId", "Total Camps"]
    })
)

# Order template
order_template = [
    "Doctor",
    "Doctor ID",
    "Doc Total Camps",
    "Total Tests",
    "Total Rx",
    "Total Strips"
]

# Start ordering
ordered_columns = ["empId", "Total Camps"]

# Append each doctor's metrics
for idx in doctor_indices:
    for field in order_template:
        col_name = f"{field} {idx}"
        if col_name in pivoted.columns:
            ordered_columns.append(col_name)

pivoted = pivoted.reindex(columns=ordered_columns)


# ------------------------------------------------------------
# 5️⃣ JOIN WITH EMPLOYEE MASTER
# ------------------------------------------------------------
final_df = emp_df.merge(pivoted, on="empId", how="left")

# Move empId first
cols = final_df.columns.tolist()
cols = ['empId'] + [c for c in cols if c != "empId"]
final_df = final_df[cols]

final_df = final_df.drop(columns=["region_list"], errors="ignore")

# ------------------------------------------------------------
# 6️⃣ EXPORT TO EXCEL
# ------------------------------------------------------------
output_file = "lupin_hb_employee_doctor_report.xlsx"
final_df.to_excel(output_file, index=False)

print("✅ Final Report Generated:", output_file)









# ------------------------------------------------------------
# 7️⃣ CREATE SUMMARY SHEET (MR → ABM → RBM → SM)
# CLEAN + READABLE VERSION
# ------------------------------------------------------------

# 1. Extract MR-level total camps
mr_camps = final_df[["empId", "Total Camps"]].copy()

# 2. Attach hierarchy columns
mr_camps = mr_camps.merge(
    emp_df[[
        "empId", "mr_name", "abm_name", "rbm_name", "sm_name",
        "state", "city", "mr_designation", "hq"
    ]],
    on="empId", how="left"
)

# ------------------------------------------------------------
# 3️⃣ Build MR rows (the only real camp contributors)
# ------------------------------------------------------------
mr_rows = mr_camps.assign(
    name=mr_camps["mr_name"],
    designation="mr",
    total_camps=mr_camps["Total Camps"].fillna(0).astype(float),
    expected_camps=2.0
)[[
    "empId", "name", "abm_name", "rbm_name", "sm_name",
    "state", "city", "designation", "hq",
    "total_camps", "expected_camps"
]]

# ------------------------------------------------------------
# 4️⃣ Build ABM/RBM/SM rows with zero placeholders
# ------------------------------------------------------------
manager_roles = ["abm", "rbm", "sm"]
manager_rows = []

for role in manager_roles:
    subset = emp_df[emp_df["mr_designation"].str.lower() == role]
    subset = subset.rename(columns={"mr_name": "name"})[
        ["empId", "name", "abm_name", "rbm_name", "sm_name", "state", "city", "hq"]
    ]
    subset["designation"] = role
    subset["total_camps"] = 0.0
    subset["expected_camps"] = 0.0
    manager_rows.append(subset)

manager_df = pd.concat(manager_rows, ignore_index=True)

# Combine MR + manager records
summary_df = pd.concat([mr_rows, manager_df], ignore_index=True)


all_sms = emp_df["sm_name"].dropna().unique().tolist()
all_rbms = emp_df["rbm_name"].dropna().unique().tolist()
all_abms = emp_df["abm_name"].dropna().unique().tolist()

new_rows = []

for sm in all_sms:
    exists = ((summary_df["designation"] == "sm") & 
              (summary_df["name"] == sm)).any()
    if not exists:
        new_rows.append({
            "empId": "",
            "name": sm,
            "designation": "sm",
            "mr_name": sm,
            "abm_name": sm,
            "rbm_name": sm,
            "sm_name": sm,
            "state": "",
            "city": "",
            "hq": "",
            "total_camps": 0.0,
            "expected_camps": 0.0,
        })

for rbm in all_rbms:
    exists = ((summary_df["designation"] == "rbm") & 
              (summary_df["name"] == rbm)).any()
    if not exists:
        new_rows.append({
            "empId": "",
            "name": rbm,
            "designation": "rbm",
            "mr_name": rbm,
            "abm_name": rbm,
            "rbm_name": rbm,
            "sm_name": rbm,   # temporary, will get corrected later
            "state": "",
            "city": "",
            "hq": "",
            "total_camps": 0.0,
            "expected_camps": 0.0,
        })

for abm in all_abms:
    exists = ((summary_df["designation"] == "abm") &
              (summary_df["name"] == abm)).any()
    if not exists:
        new_rows.append({
            "empId": "",
            "name": abm,
            "designation": "abm",
            "mr_name": abm,
            "abm_name": abm,
            "rbm_name": abm,   # temporary, corrected later
            "sm_name": abm,    # temporary
            "state": "",
            "city": "",
            "hq": "",
            "total_camps": 0.0,
            "expected_camps": 0.0,
        })

if new_rows:
    summary_df = pd.concat([summary_df, pd.DataFrame(new_rows)], ignore_index=True)



# ------------------------------------------------------------
# 5️⃣ Remove duplicates — keep highest designation
# ------------------------------------------------------------
designation_rank = {"sm": 4, "rbm": 3, "abm": 2, "mr": 1}

summary_df["designation"] = summary_df["designation"].str.lower()
summary_df["rank"] = summary_df["designation"].map(designation_rank)

summary_df = summary_df.sort_values("rank", ascending=False).drop_duplicates(subset=["empId"], keep="first")


# ------------------------------------------------------------
# 6️⃣ TRUE MR dataset (for hierarchical aggregation)
# ------------------------------------------------------------
true_mr = emp_df[emp_df["mr_designation"].str.lower() == "mr"][[
    "empId", "mr_name", "abm_name", "rbm_name", "sm_name"
]]

true_mr = true_mr.merge(
    mr_camps[["empId", "Total Camps"]],
    on="empId", how="left"
)

true_mr["total_camps"] = true_mr["Total Camps"].fillna(0)



# ------------------------------------------------------------
# ⭐ 0) FIX ABM NAMES PER RBM BLOCK (NEW RULE)
# ------------------------------------------------------------

df = summary_df.copy()

for abm, abm_group in df.groupby("abm_name", dropna=False):

    # Find TRUE ABM block → where MR name matches ABM name
    true_block = abm_group[abm_group["name"] == abm]

    if true_block.empty:
        continue  # no matching MR = skip

    true_rbm = true_block.iloc[0]["rbm_name"]  # the rbm where ABM is true

    # All RBMs under this ABM
    rbm_list = abm_group["rbm_name"].fillna("").unique().tolist()

    for rbm in rbm_list:

        # TRUE RBM block → do nothing
        if rbm == true_rbm:
            continue

        # Other RBM blocks → ABM becomes Vacant(RBM)
        vacant_abm = f"Vacant ({rbm})"

        mask = (df["abm_name"] == abm) & (df["rbm_name"] == rbm)

        # Update ABM name
        df.loc[mask, "abm_name"] = vacant_abm
        
        # Update "name" only for ABM designation row
        df.loc[mask & (df["designation"] == "abm"), "name"] = vacant_abm

summary_df = df.copy()


# ------------------------------------------------------------
# ⭐ 1) FIX RBM NAMES PER SM BLOCK (Opposite Rule)
# ------------------------------------------------------------

df = summary_df.copy()

for rbm, rbm_group in df.groupby("rbm_name", dropna=False):

    # Identify TRUE RBM block → where the RBM row exists
    rbm_row = rbm_group[(rbm_group["designation"] == "rbm") & 
                        (rbm_group["name"] == rbm)]
    
    if rbm_row.empty:
        continue  # No RBM row found (should not happen)

    true_sm = rbm_row.iloc[0]["sm_name"]

    # SMs under this RBM
    all_sms = rbm_group["sm_name"].fillna("").unique().tolist()

    for sm in all_sms:
        # All NOT TRUE blocks should become Vacant RBMs
        if sm != true_sm:
            vacant_name = f"Vacant ({sm})"
            mask = (df["rbm_name"] == rbm) & (df["sm_name"] == sm)
            df.loc[mask, "rbm_name"] = vacant_name

summary_df = df.copy()


# ------------------------------------------------------------
# 🔄 RECOMPUTE group_maps AFTER renaming (SMALL FIX)
# ------------------------------------------------------------
true_mr_after = summary_df[summary_df["designation"] == "mr"].copy()

group_maps = {}

for role, col in [("abm", "abm_name"), ("rbm", "rbm_name"), ("sm", "sm_name")]:
    g = true_mr_after.groupby(col)["total_camps"].agg(["sum", "count"]).reset_index()
    g["expected"] = g["count"] * 2

    group_maps[role] = {
        "total": dict(zip(g[col], g["sum"])),
        "expected": dict(zip(g[col], g["expected"]))
    }


# ------------------------------------------------------------
# 8️⃣ Apply the aggregated totals to summary_df
# ------------------------------------------------------------
for role in manager_roles:
    mask = summary_df["designation"] == role
    name_col = "name"

    summary_df.loc[mask, "total_camps"] = (
        summary_df.loc[mask, name_col].map(group_maps[role]["total"]).fillna(0)
    )

    summary_df.loc[mask, "expected_camps"] = (
        summary_df.loc[mask, name_col].map(group_maps[role]["expected"]).fillna(0)
    )

# ------------------------------------------------------------
# 9️⃣ Execution % (safe division)
# ------------------------------------------------------------
summary_df["execution_percent"] = summary_df.apply(
    lambda r: (r["total_camps"] / r["expected_camps"] * 100)
    if r["expected_camps"] else 0,
    axis=1
).round(0).astype(int)


# ------------------------------------------------------------
# ⭐ 2) BUILD WATERFALL SUMMARY (MR → ABM → RBM → SM)
# ------------------------------------------------------------

final_rows = []
assigned_mr_indices = set()   # prevent MRs from being counted twice

for sm, sm_group in summary_df.groupby("sm_name", dropna=False):

    for rbm, rbm_group in sm_group.groupby("rbm_name", dropna=False):

        rbm_str = str(rbm) if not pd.isna(rbm) else ""
        is_vacant_rbm = rbm_str.startswith("Vacant (")

        # ---------------------------------------------------------
        # Identify REAL ABMs only (exclude Vacant(...))
        # ---------------------------------------------------------
        real_abm_groups = {}

        for abm, abm_group in rbm_group.groupby("abm_name", dropna=False):

            if isinstance(abm, str) and abm.startswith("Vacant ("):
                continue

            if abm_group[abm_group["designation"] == "abm"].shape[0] == 0:
                continue

            real_abm_groups[abm] = abm_group

        # ---------------------------------------------------------
        # 1️⃣ REAL ABM BLOCK → Add MRs + ABM row
        # ---------------------------------------------------------
        mrs_assigned_local = set()

        for abm, abm_group in real_abm_groups.items():

            mrs = abm_group[abm_group["designation"] == "mr"]

            for _, row in mrs.iterrows():

                if row.name not in assigned_mr_indices:
                    final_rows.append(row)
                    assigned_mr_indices.add(row.name)
                    mrs_assigned_local.add(row.name)

            abm_row = abm_group[abm_group["designation"] == "abm"]
            if not abm_row.empty:
                final_rows.append(abm_row.iloc[0])

        # ---------------------------------------------------------
        # 2️⃣ LEFTOVER MRs (not under any real ABM)
        # ---------------------------------------------------------
        all_mrs = rbm_group[rbm_group["designation"] == "mr"]
        leftover_mrs = all_mrs[~all_mrs.index.isin(mrs_assigned_local)]

        leftover_mrs_unassigned = leftover_mrs[~leftover_mrs.index.isin(assigned_mr_indices)]

        for _, row in leftover_mrs_unassigned.iterrows():
            final_rows.append(row)
            assigned_mr_indices.add(row.name)

        # ---------------------------------------------------------
        # 3️⃣ VACANT ABM BLOCK
        # ---------------------------------------------------------
        if not leftover_mrs_unassigned.empty:

            total_camps = leftover_mrs_unassigned["total_camps"].sum()
            expected_camps = leftover_mrs_unassigned["expected_camps"].sum()

            exec_percent = (
                round((total_camps / expected_camps) * 100)
                if expected_camps else 0
            )

            vacant_abm = {
                "empId": "",
                "name": f"Vacant ({rbm})",
                "abm_name": f"Vacant ({rbm})",
                "rbm_name": rbm,
                "sm_name": sm,
                "state": "",
                "city": "",
                "designation": "abm",
                "hq": "",
                "total_camps": total_camps,
                "expected_camps": expected_camps,
                "execution_percent": exec_percent
            }

            final_rows.append(pd.Series(vacant_abm))

        # ---------------------------------------------------------
        # 4️⃣ RBM ROW (real or synthetic)
        # ---------------------------------------------------------
        rbm_row = rbm_group[rbm_group["designation"] == "rbm"]

        if not is_vacant_rbm:
            if not rbm_row.empty:
                final_rows.append(rbm_row.iloc[0])

        else:
            abm_rows = rbm_group[rbm_group["designation"] == "abm"]

            total_camps = abm_rows["total_camps"].sum()
            expected_camps = abm_rows["expected_camps"].sum()

            exec_percent = (
                round((total_camps / expected_camps) * 100)
                if expected_camps else 0
            )

            synthetic_rbm = {
                "empId": "",
                "name": rbm,
                "abm_name": rbm,
                "rbm_name": rbm,
                "sm_name": sm,
                "state": "",
                "city": "",
                "designation": "rbm",
                "hq": "",
                "total_camps": total_camps,
                "expected_camps": expected_camps,
                "execution_percent": exec_percent
            }

            final_rows.append(pd.Series(synthetic_rbm))

    # ---------------------------------------------------------
    # 5️⃣ SM ROW
    # ---------------------------------------------------------
    sm_row = sm_group[sm_group["designation"] == "sm"]
    if not sm_row.empty:
        final_rows.append(sm_row.iloc[0])


waterfall_df = pd.DataFrame(final_rows)

# ------------------------------------------------------------
# ⭐ RECOMPUTE TOTALS AFTER CORRECT GROUPING (IMPORTANT)
# ------------------------------------------------------------

wf = waterfall_df.copy()
wf["designation"] = wf["designation"].str.lower()

# Only MR rows contribute to totals
mr_only = wf[wf["designation"] == "mr"]

# Recompute ABM totals
abm_totals = mr_only.groupby("abm_name")[["total_camps", "expected_camps"]].sum()

# Recompute RBM totals
rbm_totals = mr_only.groupby("rbm_name")[["total_camps", "expected_camps"]].sum()

# Recompute SM totals
sm_totals = mr_only.groupby("sm_name")[["total_camps", "expected_camps"]].sum()

def apply_totals(row):
    n = row["name"]

    if row["designation"] == "abm" and n in abm_totals.index:
        row["total_camps"] = abm_totals.loc[n, "total_camps"]
        row["expected_camps"] = abm_totals.loc[n, "expected_camps"]

    elif row["designation"] == "rbm" and n in rbm_totals.index:
        row["total_camps"] = rbm_totals.loc[n, "total_camps"]
        row["expected_camps"] = rbm_totals.loc[n, "expected_camps"]

    elif row["designation"] == "sm" and n in sm_totals.index:
        row["total_camps"] = sm_totals.loc[n, "total_camps"]
        row["expected_camps"] = sm_totals.loc[n, "expected_camps"]

    return row

waterfall_df = wf.apply(apply_totals, axis=1)

# Execution %
waterfall_df["execution_percent"] = waterfall_df.apply(
    lambda r: (r["total_camps"] / r["expected_camps"] * 100)
    if r["expected_camps"] else 0,
    axis=1
).round(0).astype(int)



# ------------------------------------------------------------
# ⭐ 3) EXPORT WATERFALL SUMMARY
# ------------------------------------------------------------
with pd.ExcelWriter(
    "lupin_hb_employee_doctor_report.xlsx",
    engine="openpyxl",
    mode="a"
) as writer:
    waterfall_df.to_excel(writer, sheet_name="Waterfall Summary", index=False)

print("Waterfall Summary sheet built:", waterfall_df.shape)



# ------------------------------------------------------------
# ⭐ 4) RBM-ONLY SUMMARY SHEET
# ------------------------------------------------------------

rbm_only = summary_df[summary_df["designation"] == "rbm"].copy()

with pd.ExcelWriter(
    "lupin_hb_employee_doctor_report.xlsx",
    engine="openpyxl",
    mode="a"
) as writer:
    rbm_only.to_excel(writer, sheet_name="RBM Summary", index=False)


# ------------------------------------------------------------
# ⭐ TESTS DONE BY NON-MRs (from final_df, cleaned)
# ------------------------------------------------------------

# 1️⃣ Get NON-MR employees from emp_df
non_mr = emp_df[emp_df["mr_designation"].str.lower() != "mr"].copy()
non_mr["empId"] = non_mr["empId"].astype(str)

# 2️⃣ Merge with final_df to get activity columns
non_mr_activity = non_mr.merge(
    final_df[["empId", "Total Camps"] + 
             [col for col in final_df.columns if col.startswith("Total Tests")]],
    on="empId",
    how="left"
)

# 3️⃣ Fill NaN with zero
test_cols = [col for col in non_mr_activity.columns if col.startswith("Total Tests")]
non_mr_activity[test_cols] = non_mr_activity[test_cols].fillna(0)
non_mr_activity["Total Camps"] = non_mr_activity["Total Camps"].fillna(0)

# 4️⃣ Compute overall totals
non_mr_activity["overall_total_tests"] = non_mr_activity[test_cols].sum(axis=1)
non_mr_activity["overall_total_camps"] = non_mr_activity["Total Camps"]

# 5️⃣ Keep only active non-MRs
non_mr_activity = non_mr_activity[
    (non_mr_activity["overall_total_tests"] > 0) |
    (non_mr_activity["overall_total_camps"] > 0)
].copy()

# 6️⃣ Select only minimal columns
non_mr_activity = non_mr_activity[
    ["empId", "mr_name", "mr_designation", "hq",
     "overall_total_tests", "overall_total_camps"]
]

# 7️⃣ Export sheet
with pd.ExcelWriter(
    "lupin_hb_employee_doctor_report.xlsx",
    engine="openpyxl",
    mode="a"
) as writer:
    non_mr_activity.to_excel(writer, sheet_name="Non MR Tests", index=False)

print("Non-MR Tests sheet built:", non_mr_activity.shape)













# ------------------------------------------------------------
# 9️⃣ CREATE ONE SHEET PER RBM WITH FULL HIERARCHY (MR → ABM → RBM)
# ------------------------------------------------------------

# UNIQUE RBM NAMES ONLY
rbm_list = waterfall_df["rbm_name"].fillna("").unique().tolist()

# Track to ensure unique sheet names
used_sheet_names = set()
blank_count = 0

with pd.ExcelWriter("lupin_hb_employee_doctor_report.xlsx", engine="openpyxl", mode="a") as writer:

    for rbm in rbm_list:

        # raw RBM name (may be blank)
        raw_name = rbm if isinstance(rbm, str) else ""

        # FILTER: all rows under this RBM only
        rbm_block = waterfall_df[
            (waterfall_df["rbm_name"].fillna("") == raw_name) &
            (waterfall_df["designation"].isin(["mr", "abm", "rbm"]))
        ].copy()

        # if no rows, skip
        if rbm_block.empty:
            continue

        # BUILD EXCEL-SAFE SHEET NAME
        if raw_name.strip() == "":
            # blank RBM → create sheet with spaces
            blank_count += 1
            safe_sheet_name = " " * blank_count
        else:
            safe_sheet_name = raw_name.strip()[:31]
            for bad, rep in [("/", "-"), ("\\", "-"), ("[", "("), ("]", ")"),
                             ("?", ""), ("*", ""), (":", "-")]:
                safe_sheet_name = safe_sheet_name.replace(bad, rep)

        # Ensure name is unique
        original = safe_sheet_name
        counter = 1
        while safe_sheet_name in used_sheet_names:
            suffix = f"_{counter}"
            safe_sheet_name = original[:31 - len(suffix)] + suffix
            counter += 1

        used_sheet_names.add(safe_sheet_name)

        # WRITE SHEET
        rbm_block.to_excel(writer, sheet_name=safe_sheet_name, index=False)

        print(f"Created RBM Sheet: '{safe_sheet_name}'  (RBM: '{raw_name}')")



# ------------------------------------------------------------
# 🔟 APPLY PROFESSIONAL EXCEL STYLING (CUSTOM COLORS)
# ------------------------------------------------------------
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# Load workbook
wb = load_workbook("lupin_hb_employee_doctor_report.xlsx")

# Color codes
COLOR_HEADER = "9DC3E6"   # Blue header
COLOR_MR     = "FFFFFF"   # White
COLOR_ABM    = "FFF7CE"   # Light Yellow
COLOR_RBM    = "F8CBAD"   # Light Red
COLOR_SM     = "E4DFEC"   # Light Purple

# Border style
thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

# Apply styling sheet-by-sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    header = True

    # Identify designation column
    header_row = [cell.value for cell in ws[1]]
    if "designation" in [str(h).lower() for h in header_row]:
        desig_col = header_row.index("designation") + 1
    else:
        desig_col = None

    for row in ws.iter_rows():
        for cell in row:

            # Auto column width
            col_letter = cell.column_letter
            cell_len = len(str(cell.value)) if cell.value else 0
            ws.column_dimensions[col_letter].width = max(
                ws.column_dimensions[col_letter].width or 10,
                cell_len + 2
            )

            # ----- HEADER STYLING -----
            if header:
                cell.font = Font(bold=True, color="000000")
                cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            # ----- DESIGNATION-BASED ROW COLORING -----
            if desig_col:
                desig_value = ws.cell(row=cell.row, column=desig_col).value
                d = str(desig_value).strip().lower() if desig_value else ""

                if d == "mr":
                    fill_color = COLOR_MR
                    bold = False
                elif d == "abm":
                    fill_color = COLOR_ABM
                    bold = True
                elif d == "rbm":
                    fill_color = COLOR_RBM
                    bold = True
                elif d == "sm":
                    fill_color = COLOR_SM
                    bold = True
                else:
                    fill_color = COLOR_MR
                    bold = False

                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(bold=bold)

            cell.border = border
            cell.alignment = Alignment(vertical="center")

        header = False

# Save workbook
wb.save("lupin_hb_employee_doctor_report.xlsx")

print("🎨 Excel Styling Applied Successfully!")
