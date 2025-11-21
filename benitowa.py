import pyodbc
import pandas as pd
from google.cloud import firestore, storage
from datetime import datetime, date, timedelta
import json
import re
import warnings
warnings.filterwarnings('ignore')

# ------------------------------------------------------------
# 1️⃣ LOAD EMPLOYEE MASTER FROM FIREBASE
# ------------------------------------------------------------
st = storage.Client("neodocs-8d6cd")
bucket = st.bucket("neodocs-8d6cd-utils")

# ⚠️ Update this blob path if Benitowa uses a different file
blob = bucket.blob("org_access_codes/benitowa-uacr.json")
json_data = blob.download_as_text()
employee_dict = json.loads(json_data)

# Convert to DataFrame
emp_df = pd.DataFrame(employee_dict).transpose().reset_index()
emp_df.rename(columns={"index": "empId", "mr_region": "hq"}, inplace=True)

# Remove Training rows
emp_df = emp_df[emp_df['mr_name'] != "Training"]

# Convert region_list to comma-separated text
if "region_list" in emp_df.columns:
    emp_df["region_list"] = emp_df["region_list"].apply(
        lambda x: ", ".join(x) if isinstance(x, list) else x
    )

# Ensure empId is string
emp_df["empId"] = emp_df["empId"].astype(str)

print("Employee Master Loaded:", emp_df.shape)

# ------------------------------------------------------------
# 2️⃣ SQL SERVER — LOAD DOCTOR METRICS
# ------------------------------------------------------------
server = 'neodocs-sql-server.database.windows.net'
database = 'neodocs-sql-db'
username = 'ndDashboard'
password = 'NeoDocs@2025'
driver = '{ODBC Driver 18 for SQL Server}'

conn_str = f"""
DRIVER={driver};
SERVER={server};
DATABASE={database};
UID={username};
PWD={password};
Encrypt=yes;
TrustServerCertificate=no;
Connection Timeout=30;
"""

conn = pyodbc.connect(conn_str)
print("Connected to SQL Server")

# ------------------------------------------------------------
# DATE RANGE — CURRENT MONTH
# ------------------------------------------------------------
today = date.today()

# First day of current month
first_day_this_month = today.replace(day=1)

# Start = first day of current month
START_DATE = first_day_this_month.strftime('%Y-%m-%d')

# End = today (you can change to end of month if needed)
END_DATE = today.strftime('%Y-%m-%d')

print("SQL Range (Current Month):", START_DATE, "to", END_DATE)

# ------------------------------------------------------------
# 3️⃣ CLEAN SQL QUERY (Same structure, different brand if needed)
#    ⚠️ Update a.aId for Benitowa if required
# ------------------------------------------------------------
query = """
WITH test_summary AS (
    SELECT 
        u.empId,
        u.docId AS [Doctor ID],
        u.drName AS Doctor,
        u.oId,
        COUNT(DISTINCT u.testId) AS [Total Tests],
        COUNT(DISTINCT CONVERT(date, u.campDate)) AS [Total Camps]
    FROM dbo.user_tests u
    LEFT JOIN dbo.aId a ON u.aId = a.aId
    WHERE 
        a.aId = '4d2cce3a-58be-4143-9674-f78f3f1c32e2'
        AND u.statusCode = 200
        AND u.isDeleted = 0
        AND u.campDate BETWEEN ? AND ?
    GROUP BY u.empId, u.docId, u.drName, u.oId
),

rx_summary AS (
    SELECT 
        r.aId,
        r.oId,
        r.campDate,
        SUM(TRY_CAST(LEFT(JSON_VALUE(r.prescriptions, '$.benitowa'), 
            CHARINDEX('|', JSON_VALUE(r.prescriptions, '$.benitowa')) - 1) AS INT)) AS [Total Rx],
        SUM(TRY_CAST(LTRIM(RIGHT(JSON_VALUE(r.prescriptions, '$.benitowa'), 
            LEN(JSON_VALUE(r.prescriptions, '$.benitowa')) - 
            CHARINDEX('|', JSON_VALUE(r.prescriptions, '$.benitowa')))) AS INT)) AS [Total Strips]
    FROM dbo.rx r
    LEFT JOIN dbo.aId a ON r.aId = a.aId
    WHERE 
        a.aId = '4d2cce3a-58be-4143-9674-f78f3f1c32e2'
        AND r.isDeleted = 0
        AND r.campDate BETWEEN ? AND ?
    GROUP BY r.aId, r.oId, r.campDate
)

SELECT 
    t.empId AS [empId],
    t.Doctor,
    t.[Doctor ID],
    t.[Total Camps] AS [Doc Total Camps],
    ISNULL(SUM(rx.[Total Rx]), 0) AS [Total Rx],
    ISNULL(SUM(rx.[Total Strips]), 0) AS [Total Strips],
    t.[Total Tests]
FROM test_summary t
LEFT JOIN rx_summary rx
    ON t.oId = rx.oId
GROUP BY 
    t.empId, t.Doctor, t.[Doctor ID], t.[Total Camps], t.[Total Tests]
ORDER BY t.empId;
"""

params = [START_DATE, END_DATE, START_DATE, END_DATE]
sql_df = pd.read_sql(query, conn, params=params)

print("Doctor Metrics Loaded:", sql_df.shape)

# ------------------------------------------------------------
# 4️⃣ ASSIGN DOCTOR INDEX PER empId
# ------------------------------------------------------------
sql_df['Doctor Index'] = sql_df.groupby(['empId']).cumcount() + 1

# ------------------------------------------------------------
# 5️⃣ PIVOT DOCTOR METRICS — Wide format
# ------------------------------------------------------------
pivoted = sql_df.pivot_table(
    index=['empId'],
    columns='Doctor Index',
    values=['Doctor', 'Doctor ID', 'Doc Total Camps', 'Total Rx', 'Total Strips', 'Total Tests'],
    aggfunc='first'
)

# Flatten columns
pivoted.columns = [f"{col[0]} {int(col[1])}" for col in pivoted.columns]
pivoted = pivoted.reset_index()

# ------------------------------------------------------------
# ADD TOTAL CAMPS COLUMN (SUM OF ALL DOC TOTAL CAMPS)
# ------------------------------------------------------------
doc_camp_cols = [col for col in pivoted.columns if col.startswith("Doc Total Camps")]
pivoted["Total Camps"] = pivoted[doc_camp_cols].sum(axis=1)

# ------------------------------------------------------------
# REORDER COLUMNS WITH TOTAL CAMPS FIRST, THEN DOCTORS
# ------------------------------------------------------------
doctor_indices = sorted(
    list({
        int(col.split()[-1]) 
        for col in pivoted.columns 
        if col not in ["empId", "Total Camps"]
    })
)

order_template = [
    "Doctor",
    "Doctor ID",
    "Doc Total Camps",
    "Total Tests",
    "Total Rx",
    "Total Strips"
]

ordered_columns = ["empId", "Total Camps"]
for idx in doctor_indices:
    for field in order_template:
        col_name = f"{field} {idx}"
        if col_name in pivoted.columns:
            ordered_columns.append(col_name)

pivoted = pivoted.reindex(columns=ordered_columns)

# ------------------------------------------------------------
# 6️⃣ JOIN WITH EMPLOYEE MASTER → FINAL EMPLOYEE-DOCTOR TABLE
# ------------------------------------------------------------
final_df = emp_df.merge(pivoted, on="empId", how="left")

# Move empId first
cols = final_df.columns.tolist()
cols = ['empId'] + [c for c in cols if c != "empId"]
final_df = final_df[cols]

final_df = final_df.drop(columns=["region_list"], errors="ignore")

# ------------------------------------------------------------
# 7️⃣ EXPORT BASE REPORT (EMPLOYEE + DOCTORS)
# ------------------------------------------------------------
output_file = "benitowa_employee_doctor_report.xlsx"
final_df.to_excel(output_file, index=False)
print("✅ Base Report Generated:", output_file)

# ============================================================
# 🔁 BENITOWA HIERARCHY SUMMARY (ABM → RBM → SM)
#    🔹 Worker level = ABM (no MR)
#    🔹 Expected camps per ABM = 2
#    🔹 RBM & SM = rolled up from ABMs
# ============================================================

# ------------------------------------------------------------
# 8️⃣ Build ABM-level activity (worker layer)
# ------------------------------------------------------------
# Extract Total Camps from final_df and attach hierarchy info
abm_camps = final_df[["empId", "Total Camps"]].copy()
abm_camps = abm_camps.merge(
    emp_df[[
        "empId", "mr_name", "abm_name", "rbm_name", "sm_name",
        "state", "city", "mr_designation", "hq"
    ]],
    on="empId", how="left"
)

# Normalize designation
abm_camps["designation"] = abm_camps["mr_designation"].str.lower().fillna("")

# ABM = worker level (like MR earlier)
abm_rows = abm_camps[abm_camps["designation"] == "abm"].copy()
abm_rows["name"] = abm_rows["mr_name"]
abm_rows["total_camps"] = abm_rows["Total Camps"].fillna(0).astype(float)
abm_rows["expected_camps"] = 4.0  # ⭐ Benitowa: 2 camps per ABM

abm_rows = abm_rows[[
    "empId", "name", "abm_name", "rbm_name", "sm_name",
    "state", "city", "designation", "hq",
    "total_camps", "expected_camps"
]]

# ------------------------------------------------------------
# 9️⃣ Manager rows (RBM & SM) with zero placeholders
# ------------------------------------------------------------
manager_roles = ["rbm", "sm"]
manager_rows = []

for role in manager_roles:
    subset = emp_df[emp_df["mr_designation"].str.lower() == role].copy()
    if subset.empty:
        continue

    subset["name"] = subset["mr_name"]
    subset["designation"] = role
    subset["total_camps"] = 0.0
    subset["expected_camps"] = 0.0

    subset = subset[[
        "empId", "name", "abm_name", "rbm_name", "sm_name",
        "state", "city", "designation", "hq",
        "total_camps", "expected_camps"
    ]]
    manager_rows.append(subset)

if manager_rows:
    manager_df = pd.concat(manager_rows, ignore_index=True)
else:
    manager_df = pd.DataFrame(columns=abm_rows.columns)

# Combine ABM + manager records
summary_df = pd.concat([abm_rows, manager_df], ignore_index=True)

# ------------------------------------------------------------
# 🔟 Remove duplicates — keep highest designation
#     (SM > RBM > ABM)
# ------------------------------------------------------------
designation_rank = {"sm": 3, "rbm": 2, "abm": 1}
summary_df["rank"] = summary_df["designation"].map(designation_rank).fillna(0)
summary_df = summary_df.sort_values("rank", ascending=False).drop_duplicates(
    subset=["empId"], keep="first"
)

# ------------------------------------------------------------
# 1️⃣1️⃣ Compute roll-ups for RBM & SM from ABMs
# ------------------------------------------------------------
worker = abm_rows.copy()
worker["total_camps"] = worker["total_camps"].fillna(0)
worker["expected_camps"] = worker["expected_camps"].fillna(0)

group_maps = {}

# RBM aggregation
g_rbm = worker.groupby("rbm_name", dropna=False)[["total_camps", "expected_camps"]].sum().reset_index()
group_maps["rbm"] = {
    "total": dict(zip(g_rbm["rbm_name"], g_rbm["total_camps"])),
    "expected": dict(zip(g_rbm["rbm_name"], g_rbm["expected_camps"]))
}

# SM aggregation
g_sm = worker.groupby("sm_name", dropna=False)[["total_camps", "expected_camps"]].sum().reset_index()
group_maps["sm"] = {
    "total": dict(zip(g_sm["sm_name"], g_sm["total_camps"])),
    "expected": dict(zip(g_sm["sm_name"], g_sm["expected_camps"]))
}

# Apply to summary_df manager rows
for role in manager_roles:
    if role == "rbm":
        key_col = "rbm_name"
    elif role == "sm":
        key_col = "sm_name"
    else:
        continue

    mask = summary_df["designation"] == role
    summary_df.loc[mask, "total_camps"] = summary_df.loc[mask, key_col].map(
        group_maps[role]["total"]
    ).fillna(0)

    summary_df.loc[mask, "expected_camps"] = summary_df.loc[mask, key_col].map(
        group_maps[role]["expected"]
    ).fillna(0)

# ------------------------------------------------------------
# 1️⃣2️⃣ Execution % (safe division)
# ------------------------------------------------------------
summary_df["execution_percent"] = summary_df.apply(
    lambda r: (r["total_camps"] / r["expected_camps"] * 100)
    if r["expected_camps"] else 0,
    axis=1
).round(0).astype(int)

# ------------------------------------------------------------
# 1️⃣3️⃣ BUILD WATERFALL SUMMARY (ABM → RBM → SM)
# ------------------------------------------------------------
final_rows = []

# Group by SM
for sm, sm_group in summary_df.groupby("sm_name", dropna=False):

    # All RBMs under this SM
    rbm_list = sm_group["rbm_name"].fillna("").unique().tolist()

    for rbm in rbm_list:
        raw_rbm = rbm if isinstance(rbm, str) else ""

        rbm_block = summary_df[
            (summary_df["sm_name"] == sm) &
            (summary_df["rbm_name"].fillna("") == raw_rbm)
        ].copy()

        if rbm_block.empty:
            continue

        # 1) ABMs under this RBM
        abm_block = rbm_block[rbm_block["designation"] == "abm"].copy()

        for _, row in abm_block.sort_values("name").iterrows():
            final_rows.append(row)

        # 2) RBM row
        rbm_row = rbm_block[rbm_block["designation"] == "rbm"]
        if not rbm_row.empty:
            final_rows.append(rbm_row.iloc[0])

    # 3) SM row (one per SM)
    sm_row = sm_group[sm_group["designation"] == "sm"]
    if not sm_row.empty:
        final_rows.append(sm_row.iloc[0])

waterfall_df = pd.DataFrame(final_rows).reset_index(drop=True)
waterfall_df = waterfall_df.drop(columns=["rank"], errors="ignore")

# ------------------------------------------------------------
# 1️⃣4️⃣ EXPORT WATERFALL SUMMARY
# ------------------------------------------------------------
with pd.ExcelWriter(
    output_file,
    engine="openpyxl",
    mode="a"
) as writer:
    waterfall_df.to_excel(writer, sheet_name="Waterfall Summary", index=False)

print("Waterfall Summary sheet built:", waterfall_df.shape)

# ------------------------------------------------------------
# 1️⃣5️⃣ RBM-ONLY SUMMARY SHEET
# ------------------------------------------------------------
rbm_only = summary_df[summary_df["designation"] == "rbm"].copy()

with pd.ExcelWriter(
    output_file,
    engine="openpyxl",
    mode="a"
) as writer:
    rbm_only.to_excel(writer, sheet_name="RBM Summary", index=False)

print("RBM Summary sheet built:", rbm_only.shape)

# ------------------------------------------------------------
# 1️⃣6️⃣ TESTS DONE BY NON-ABMs (RBM/SM etc.)
# ------------------------------------------------------------
# 1) Get NON-ABM employees from emp_df
non_abm = emp_df[emp_df["mr_designation"].str.lower() != "abm"].copy()
non_abm["empId"] = non_abm["empId"].astype(str)

# 2) Merge with final_df to get activity columns
non_abm_activity = non_abm.merge(
    final_df[["empId", "Total Camps"] +
             [col for col in final_df.columns if col.startswith("Total Tests")]],
    on="empId",
    how="left"
)

# 3) Fill NaN with zero
test_cols = [col for col in non_abm_activity.columns if col.startswith("Total Tests")]
non_abm_activity[test_cols] = non_abm_activity[test_cols].fillna(0)
non_abm_activity["Total Camps"] = non_abm_activity["Total Camps"].fillna(0)

# 4) Compute overall totals
non_abm_activity["overall_total_tests"] = non_abm_activity[test_cols].sum(axis=1)
non_abm_activity["overall_total_camps"] = non_abm_activity["Total Camps"]

# 5) Keep only active non-ABMs
non_abm_activity = non_abm_activity[
    (non_abm_activity["overall_total_tests"] > 0) |
    (non_abm_activity["overall_total_camps"] > 0)
].copy()

# 6) Select only minimal columns
non_abm_activity = non_abm_activity[
    ["empId", "mr_name", "mr_designation", "hq",
     "overall_total_tests", "overall_total_camps"]
]

# 7) Export sheet
with pd.ExcelWriter(
    output_file,
    engine="openpyxl",
    mode="a"
) as writer:
    non_abm_activity.to_excel(writer, sheet_name="Non ABM Tests", index=False)

print("Non-ABM Tests sheet built:", non_abm_activity.shape)

# ------------------------------------------------------------
# 1️⃣7️⃣ CREATE ONE SHEET PER RBM WITH FULL HIERARCHY (ABM → RBM)
# ------------------------------------------------------------
rbm_list = waterfall_df["rbm_name"].fillna("").unique().tolist()

from openpyxl import load_workbook
used_sheet_names = set()
blank_count = 0

# Ensure we don't overwrite existing sheets list
wb_tmp = load_workbook(output_file)
used_sheet_names = set(wb_tmp.sheetnames)

with pd.ExcelWriter(output_file, engine="openpyxl", mode="a") as writer:

    for rbm in rbm_list:

        raw_name = rbm if isinstance(rbm, str) else ""

        # FILTER: all rows under this RBM only (ABM + RBM)
        rbm_block = waterfall_df[
            (waterfall_df["rbm_name"].fillna("") == raw_name) &
            (waterfall_df["designation"].isin(["abm", "rbm"]))
        ].copy()

        if rbm_block.empty:
            continue

        # BUILD EXCEL-SAFE SHEET NAME
        if raw_name.strip() == "":
            blank_count += 1
            safe_sheet_name = " " * blank_count
        else:
            safe_sheet_name = raw_name.strip()[:31]
            for bad, rep in [("/", "-"), ("\\", "-"), ("[", "("), ("]", ")"),
                             ("?", ""), ("*", ""), (":", "-")]:
                safe_sheet_name = safe_sheet_name.replace(bad, rep)

        original = safe_sheet_name
        counter = 1
        while safe_sheet_name in used_sheet_names:
            suffix = f"_{counter}"
            safe_sheet_name = original[:31 - len(suffix)] + suffix
            counter += 1

        used_sheet_names.add(safe_sheet_name)

        rbm_block.to_excel(writer, sheet_name=safe_sheet_name, index=False)

        print(f"Created RBM Sheet: '{safe_sheet_name}'  (RBM: '{raw_name}')")

# ------------------------------------------------------------
# 1️⃣8️⃣ APPLY PROFESSIONAL EXCEL STYLING (CUSTOM COLORS)
# ------------------------------------------------------------
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

wb = load_workbook(output_file)

# Color codes
COLOR_HEADER = "9DC3E6"   # Blue header
COLOR_ABM    = "FFF7CE"   # Light Yellow (worker level)
COLOR_RBM    = "F8CBAD"   # Light Red
COLOR_SM     = "E4DFEC"   # Light Purple
COLOR_DEFAULT = "FFFFFF"  # White

thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    header = True

    # Identify designation column
    header_row = [cell.value for cell in ws[1]]
    header_lower = [str(h).lower() for h in header_row]
    if "designation" in header_lower:
        desig_col = header_lower.index("designation") + 1
    else:
        desig_col = None

    for row in ws.iter_rows():
        for cell in row:

            # Auto column width
            col_letter = cell.column_letter
            cell_len = len(str(cell.value)) if cell.value is not None else 0
            current_width = ws.column_dimensions[col_letter].width
            if current_width is None:
                current_width = 10
            ws.column_dimensions[col_letter].width = max(current_width, cell_len + 2)

            # ----- HEADER STYLING -----
            if header:
                cell.font = Font(bold=True, color="000000")
                cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            # ----- DESIGNATION-BASED ROW COLORING -----
            if desig_col:
                desig_value = ws.cell(row=cell.row, column=desig_col).value
                d = str(desig_value).strip().lower() if desig_value else ""

                if d == "abm":
                    fill_color = COLOR_ABM
                    bold = False
                elif d == "rbm":
                    fill_color = COLOR_RBM
                    bold = True
                elif d == "sm":
                    fill_color = COLOR_SM
                    bold = True
                else:
                    fill_color = COLOR_DEFAULT
                    bold = False

                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(bold=bold)

            cell.border = border
            cell.alignment = Alignment(vertical="center")

        header = False

wb.save(output_file)
print("🎨 Excel Styling Applied Successfully!")
print("✅ Benitowa Report Complete:", output_file)
