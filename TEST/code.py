import pyodbc
import pandas as pd
from google.cloud import firestore, storage
from datetime import datetime, date, timedelta
import json
import re
import warnings
warnings.filterwarnings('ignore')

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import math

# ------------------------------------------------------------
# 0️⃣ READ PHARMA LIST
# ------------------------------------------------------------
PHARMA_LIST_FILE = "Pharma_list.xlsx"

pharma_df = pd.read_excel(PHARMA_LIST_FILE)

# Expecting columns: pharma_name, aid, json_file, prescription_key


# ------------------------------------------------------------
# 1️⃣ COMMON SETUP: GCS + SQL + DATE RANGE
# ------------------------------------------------------------
# Firebase Storage
st = storage.Client("neodocs-8d6cd")
bucket = st.bucket("neodocs-8d6cd-utils")

# SQL SERVER — CONNECTION
server = 'neodocs-sql-server.database.windows.net'
database = 'neodocs-sql-db'
username = 'ndDashboard'
password = 'NeoDocs@2025'
driver = '{ODBC Driver 18 for SQL Server}'

conn_str = f"""
DRIVER={driver};
SERVER={server};
DATABASE={database};
UID={username};
PWD={password};
Encrypt=yes;
TrustServerCertificate=no;
Connection Timeout=30;
"""

conn = pyodbc.connect(conn_str)
print("✅ Connected to SQL Server")

# DATE RANGE — CURRENT MONTH
today = date.today()
first_day_this_month = today.replace(day=1)
START_DATE = first_day_this_month.strftime('%Y-%m-%d')
END_DATE = today.strftime('%Y-%m-%d')
print("SQL Range (Current Month):", START_DATE, "to", END_DATE)


# ------------------------------------------------------------
# 2️⃣ FUNCTION: GENERATE REPORT FOR ONE PHARMA
# ------------------------------------------------------------
def generate_pharma_report(aid, pharma_name, json_file, prescription_key,
                           conn, bucket, start_date, end_date):
    """
    Generates the full Excel report for a single pharma:
    - Reads employee master from given json_file in GCS
    - Runs SQL for given aId + prescription_key
    - Builds all sheets + styling
    """

    print(f"\n==============================")
    print(f"▶ Processing pharma: {pharma_name}")
    print(f"   aId: {aid}")
    print(f"   json_file: {json_file}")
    print(f"   prescription_key: {prescription_key}")
    print(f"==============================")

    # -----------------------------
    # CLEAN INPUTS
    # -----------------------------
    json_file = str(json_file).strip().strip("'").strip('"')
    pres_key = str(prescription_key).strip() if not (isinstance(prescription_key, float) and math.isnan(prescription_key)) else ""
    has_prescription = bool(pres_key)

    # --------------------------------------------------------
    # A) LOAD EMPLOYEE MASTER FROM FIREBASE (pharma-specific)
    # --------------------------------------------------------
    blob = bucket.blob(json_file)
    json_data = blob.download_as_text()
    employee_dict = json.loads(json_data)

    # Convert to DataFrame
    emp_df = pd.DataFrame(employee_dict).transpose().reset_index()
    emp_df.rename(columns={"index": "empId", "mr_region": "hq"}, inplace=True)

    # Remove Training rows
    emp_df = emp_df[emp_df['mr_name'] != "Training"]

    # Convert region_list to comma-separated text
    if "region_list" in emp_df.columns:
        emp_df["region_list"] = emp_df["region_list"].apply(
            lambda x: ", ".join(x) if isinstance(x, list) else x
        )

    # Ensure empId is string
    emp_df["empId"] = emp_df["empId"].astype(str)

    print("Employee Master Loaded:", emp_df.shape)

    # --------------------------------------------------------
    # B) SQL — DOCTOR METRICS (dynamic aId + prescription_key)
    # --------------------------------------------------------

    if has_prescription:
        # With Rx & Strips
        query = f"""
        WITH test_summary AS (
            SELECT 
                u.empId,
                u.docId AS [Doctor ID],
                u.drName AS Doctor,
                u.oId,
                COUNT(DISTINCT u.testId) AS [Total Tests],
                COUNT(DISTINCT CONVERT(date, u.campDate)) AS [Total Camps]
            FROM dbo.user_tests u
            LEFT JOIN dbo.aId a ON u.aId = a.aId
            WHERE 
                a.aId = '{aid}'
                AND u.statusCode = 200
                AND u.isDeleted = 0
                AND u.campDate BETWEEN ? AND ?
            GROUP BY u.empId, u.docId, u.drName, u.oId
        ),

        rx_summary AS (
            SELECT 
                r.aId,
                r.oId,
                r.campDate,
                SUM(TRY_CAST(LEFT(JSON_VALUE(r.prescriptions, '$.{pres_key}'), 
                    CHARINDEX('|', JSON_VALUE(r.prescriptions, '$.{pres_key}')) - 1) AS INT)) AS [Total Rx],
                SUM(TRY_CAST(LTRIM(RIGHT(JSON_VALUE(r.prescriptions, '$.{pres_key}'), 
                    LEN(JSON_VALUE(r.prescriptions, '$.{pres_key}')) - 
                    CHARINDEX('|', JSON_VALUE(r.prescriptions, '$.{pres_key}')))) AS INT)) AS [Total Strips]
            FROM dbo.rx r
            LEFT JOIN dbo.aId a ON r.aId = a.aId
            WHERE 
                a.aId = '{aid}'
                AND r.isDeleted = 0
                AND r.campDate BETWEEN ? AND ?
            GROUP BY r.aId, r.oId, r.campDate
        )

        SELECT 
            t.empId AS [empId],
            t.Doctor,
            t.[Doctor ID],
            t.[Total Camps] AS [Doc Total Camps],
            ISNULL(SUM(rx.[Total Rx]), 0) AS [Total Rx],
            ISNULL(SUM(rx.[Total Strips]), 0) AS [Total Strips],
            t.[Total Tests]
        FROM test_summary t
        LEFT JOIN rx_summary rx
            ON t.oId = rx.oId
        GROUP BY 
            t.empId, t.Doctor, t.[Doctor ID], t.[Total Camps], t.[Total Tests]
        ORDER BY t.empId;
        """
        params = [start_date, end_date, start_date, end_date]
    else:
        # No prescription_key -> Rx & Strips = 0
        print("No prescription_key provided → Rx/Strips will be 0 for this pharma.")
        query = f"""
        WITH test_summary AS (
            SELECT 
                u.empId,
                u.docId AS [Doctor ID],
                u.drName AS Doctor,
                u.oId,
                COUNT(DISTINCT u.testId) AS [Total Tests],
                COUNT(DISTINCT CONVERT(date, u.campDate)) AS [Total Camps]
            FROM dbo.user_tests u
            LEFT JOIN dbo.aId a ON u.aId = a.aId
            WHERE 
                a.aId = '{aid}'
                AND u.statusCode = 200
                AND u.isDeleted = 0
                AND u.campDate BETWEEN ? AND ?
            GROUP BY u.empId, u.docId, u.drName, u.oId
        )

        SELECT 
            t.empId AS [empId],
            t.Doctor,
            t.[Doctor ID],
            t.[Total Camps] AS [Doc Total Camps],
            0 AS [Total Rx],
            0 AS [Total Strips],
            t.[Total Tests]
        FROM test_summary t
        ORDER BY t.empId;
        """
        params = [start_date, end_date]

    sql_df = pd.read_sql(query, conn, params=params)
    print("Doctor Metrics Loaded:", sql_df.shape)

    # --------------------------------------------------------
    # C) ASSIGN DOCTOR INDEX PER empId
    # --------------------------------------------------------
    sql_df['Doctor Index'] = sql_df.groupby(['empId']).cumcount() + 1

    # --------------------------------------------------------
    # D) PIVOT DOCTOR METRICS — Wide format
    # --------------------------------------------------------
    pivoted = sql_df.pivot_table(
        index=['empId'],
        columns='Doctor Index',
        values=['Doctor', 'Doctor ID', 'Doc Total Camps', 'Total Rx', 'Total Strips', 'Total Tests'],
        aggfunc='first'
    )

    # Flatten columns
    pivoted.columns = [f"{col[0]} {int(col[1])}" for col in pivoted.columns]
    pivoted = pivoted.reset_index()

    # --------------------------------------------------------
    # ADD TOTAL CAMPS COLUMN (SUM OF ALL DOC TOTAL CAMPS)
    # --------------------------------------------------------
    doc_camp_cols = [col for col in pivoted.columns if col.startswith("Doc Total Camps")]
    pivoted["Total Camps"] = pivoted[doc_camp_cols].sum(axis=1)

    # --------------------------------------------------------
    # REORDER COLUMNS WITH TOTAL CAMPS FIRST, THEN DOCTORS
    # --------------------------------------------------------
    doctor_indices = sorted(
        list({
            int(col.split()[-1]) 
            for col in pivoted.columns 
            if col not in ["empId", "Total Camps"]
        })
    )

    order_template = [
        "Doctor",
        "Doctor ID",
        "Doc Total Camps",
        "Total Tests",
        "Total Rx",
        "Total Strips"
    ]

    ordered_columns = ["empId", "Total Camps"]

    for idx in doctor_indices:
        for field in order_template:
            col_name = f"{field} {idx}"
            if col_name in pivoted.columns:
                ordered_columns.append(col_name)

    pivoted = pivoted.reindex(columns=ordered_columns)

    # --------------------------------------------------------
    # E) JOIN WITH EMPLOYEE MASTER
    # --------------------------------------------------------
    final_df = emp_df.merge(pivoted, on="empId", how="left")

    cols = final_df.columns.tolist()
    cols = ['empId'] + [c for c in cols if c != "empId"]
    final_df = final_df[cols]

    final_df = final_df.drop(columns=["region_list"], errors="ignore")

    # --------------------------------------------------------
    # F) EXPORT BASE EXCEL (main sheet)
    # --------------------------------------------------------
    output_file = f"{pharma_name}_employee_doctor_report.xlsx"
    final_df.to_excel(output_file, index=False)
    print("✅ Final Report Generated:", output_file)

    # ------------------------------------------------------------
    # 7️⃣ CREATE SUMMARY SHEET (MR → ABM → RBM → SM)
    # ------------------------------------------------------------
    mr_camps = final_df[["empId", "Total Camps"]].copy()

    mr_camps = mr_camps.merge(
        emp_df[[
            "empId", "mr_name", "abm_name", "rbm_name", "sm_name",
            "state", "city", "mr_designation", "hq"
        ]],
        on="empId", how="left"
    )

    mr_rows = mr_camps.assign(
        name=mr_camps["mr_name"],
        designation="mr",
        total_camps=mr_camps["Total Camps"].fillna(0).astype(float),
        expected_camps=4.0
    )[[
        "empId", "name", "abm_name", "rbm_name", "sm_name",
        "state", "city", "designation", "hq",
        "total_camps", "expected_camps"
    ]]

    manager_roles = ["abm", "rbm", "sm"]
    manager_rows = []

    for role in manager_roles:
        subset = emp_df[emp_df["mr_designation"].str.lower() == role]
        subset = subset.rename(columns={"mr_name": "name"})[
            ["empId", "name", "abm_name", "rbm_name", "sm_name", "state", "city", "hq"]
        ]
        subset["designation"] = role
        subset["total_camps"] = 0.0
        subset["expected_camps"] = 0.0
        manager_rows.append(subset)

    manager_df = pd.concat(manager_rows, ignore_index=True)

    summary_df = pd.concat([mr_rows, manager_df], ignore_index=True)

    designation_rank = {"sm": 4, "rbm": 3, "abm": 2, "mr": 1}

    summary_df["designation"] = summary_df["designation"].str.lower()
    summary_df["rank"] = summary_df["designation"].map(designation_rank)

    summary_df = summary_df.sort_values("rank", ascending=False).drop_duplicates(subset=["empId"], keep="first")

    true_mr = emp_df[emp_df["mr_designation"].str.lower() == "mr"][[
        "empId", "mr_name", "abm_name", "rbm_name", "sm_name"
    ]]

    true_mr = true_mr.merge(
        mr_camps[["empId", "Total Camps"]],
        on="empId", how="left"
    )

    true_mr["total_camps"] = true_mr["Total Camps"].fillna(0)

    # ⭐ 0) FIX ABM NAMES PER RBM BLOCK (NEW RULE)
    df = summary_df.copy()

    for abm, abm_group in df.groupby("abm_name", dropna=False):
        true_block = abm_group[abm_group["name"] == abm]
        if true_block.empty:
            continue
        true_rbm = true_block.iloc[0]["rbm_name"]
        rbm_list_local = abm_group["rbm_name"].fillna("").unique().tolist()
        for rbm in rbm_list_local:
            if rbm == true_rbm:
                continue
            vacant_abm = f"Vacant ({rbm})"
            mask = (df["abm_name"] == abm) & (df["rbm_name"] == rbm)
            df.loc[mask, "abm_name"] = vacant_abm
            df.loc[mask & (df["designation"] == "abm"), "name"] = vacant_abm

    summary_df = df.copy()

    # ⭐ 1) FIX RBM NAMES PER SM BLOCK (Opposite Rule)
    df = summary_df.copy()

    for rbm, rbm_group in df.groupby("rbm_name", dropna=False):
        rbm_row = rbm_group[(rbm_group["designation"] == "rbm") & 
                            (rbm_group["name"] == rbm)]
        if rbm_row.empty:
            continue
        true_sm = rbm_row.iloc[0]["sm_name"]
        all_sms = rbm_group["sm_name"].fillna("").unique().tolist()
        for sm in all_sms:
            if sm != true_sm:
                vacant_name = f"Vacant ({sm})"
                mask = (df["rbm_name"] == rbm) & (df["sm_name"] == sm)
                df.loc[mask, "rbm_name"] = vacant_name

    summary_df = df.copy()

    # 🔄 RECOMPUTE group_maps AFTER renaming
    true_mr_after = summary_df[summary_df["designation"] == "mr"].copy()

    group_maps = {}

    for role, col in [("abm", "abm_name"), ("rbm", "rbm_name"), ("sm", "sm_name")]:
        g = true_mr_after.groupby(col)["total_camps"].agg(["sum", "count"]).reset_index()
        g["expected"] = g["count"] * 4

        group_maps[role] = {
            "total": dict(zip(g[col], g["sum"])),
            "expected": dict(zip(g[col], g["expected"]))
        }

    for role in manager_roles:
        mask = summary_df["designation"] == role
        name_col = "name"

        summary_df.loc[mask, "total_camps"] = (
            summary_df.loc[mask, name_col].map(group_maps[role]["total"]).fillna(0)
        )

        summary_df.loc[mask, "expected_camps"] = (
            summary_df.loc[mask, name_col].map(group_maps[role]["expected"]).fillna(0)
        )

    summary_df["execution_percent"] = summary_df.apply(
        lambda r: (r["total_camps"] / r["expected_camps"] * 100)
        if r["expected_camps"] else 0,
        axis=1
    ).round(0).astype(int)

    # ⭐ 2) BUILD WATERFALL SUMMARY (MR → ABM → RBM → SM)
    final_rows = []

    for sm, sm_group in summary_df.groupby("sm_name", dropna=False):

        for rbm, rbm_group in sm_group.groupby("rbm_name", dropna=False):

            rbm_str = str(rbm) if not pd.isna(rbm) else ""
            is_vacant_rbm = rbm_str.startswith("Vacant (")

            real_abm_groups = {}

            for abm, abm_group in rbm_group.groupby("abm_name", dropna=False):

                if isinstance(abm, str) and abm.startswith("Vacant ("):
                    continue

                if abm_group[abm_group["designation"] == "abm"].shape[0] == 0:
                    continue

                real_abm_groups[abm] = abm_group

            mrs_assigned = []

            for abm, abm_group in real_abm_groups.items():

                mrs = abm_group[abm_group["designation"] == "mr"]

                for _, row in mrs.iterrows():
                    final_rows.append(row)
                    mrs_assigned.append(row.name)

                abm_row = abm_group[abm_group["designation"] == "abm"]
                if not abm_row.empty:
                    final_rows.append(abm_row.iloc[0])

            mrs_assigned = set(mrs_assigned)

            all_mrs = rbm_group[rbm_group["designation"] == "mr"]
            leftover_mrs = all_mrs[~all_mrs.index.isin(mrs_assigned)]

            for _, row in leftover_mrs.iterrows():
                final_rows.append(row)

            if not leftover_mrs.empty:

                total_camps = leftover_mrs["total_camps"].sum()
                expected_camps = leftover_mrs["expected_camps"].sum()
                exec_percent = (
                    round((total_camps / expected_camps) * 100)
                    if expected_camps > 0 else 0
                )

                vacant_abm = {
                    "empId": "",
                    "name": f"Vacant ({rbm})",
                    "abm_name": f"Vacant ({rbm})",
                    "rbm_name": rbm,
                    "sm_name": sm,
                    "state": "",
                    "city": "",
                    "designation": "abm",
                    "hq": "",
                    "total_camps": total_camps,
                    "expected_camps": expected_camps,
                    "execution_percent": exec_percent
                }

                final_rows.append(pd.Series(vacant_abm))

            rbm_row = rbm_group[rbm_group["designation"] == "rbm"]

            if not is_vacant_rbm:
                if not rbm_row.empty:
                    final_rows.append(rbm_row.iloc[0])

            else:
                abm_rows = rbm_group[rbm_group["designation"] == "abm"]

                total_camps = abm_rows["total_camps"].sum()
                expected_camps = abm_rows["expected_camps"].sum()

                exec_percent = (
                    round((total_camps / expected_camps) * 100)
                    if expected_camps > 0 else 0
                )

                synthetic_rbm = {
                    "empId": "",
                    "name": rbm,
                    "abm_name": rbm,
                    "rbm_name": rbm,
                    "sm_name": sm,
                    "state": "",
                    "city": "",
                    "designation": "rbm",
                    "hq": "",
                    "total_camps": total_camps,
                    "expected_camps": expected_camps,
                    "execution_percent": exec_percent
                }

                final_rows.append(pd.Series(synthetic_rbm))

        sm_row = sm_group[sm_group["designation"] == "sm"]
        if not sm_row.empty:
            final_rows.append(sm_row.iloc[0])

    waterfall_df = pd.DataFrame(final_rows)
    waterfall_df = waterfall_df.drop(columns=["abm_name","sm_name","rank"], errors="ignore")

    # ⭐ 3) EXPORT WATERFALL SUMMARY
    with pd.ExcelWriter(
        output_file,
        engine="openpyxl",
        mode="a"
    ) as writer:
        waterfall_df.to_excel(writer, sheet_name="Waterfall Summary", index=False)

    print("Waterfall Summary sheet built:", waterfall_df.shape)

    # ⭐ 4) RBM-ONLY SUMMARY SHEET
    rbm_only = waterfall_df[waterfall_df["designation"] == "rbm"].copy()

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl",
        mode="a"
    ) as writer:
        rbm_only.to_excel(writer, sheet_name="RBM Summary", index=False)

    # ⭐ TESTS DONE BY NON-MRs (from final_df, cleaned)
    non_mr = emp_df[emp_df["mr_designation"].str.lower() != "mr"].copy()
    non_mr["empId"] = non_mr["empId"].astype(str)

    non_mr_activity = non_mr.merge(
        final_df[["empId", "Total Camps"] + 
                 [col for col in final_df.columns if col.startswith("Total Tests")]],
        on="empId",
        how="left"
    )

    test_cols = [col for col in non_mr_activity.columns if col.startswith("Total Tests")]
    non_mr_activity[test_cols] = non_mr_activity[test_cols].fillna(0)
    non_mr_activity["Total Camps"] = non_mr_activity["Total Camps"].fillna(0)

    non_mr_activity["overall_total_tests"] = non_mr_activity[test_cols].sum(axis=1)
    non_mr_activity["overall_total_camps"] = non_mr_activity["Total Camps"]

    non_mr_activity = non_mr_activity[
        (non_mr_activity["overall_total_tests"] > 0) |
        (non_mr_activity["overall_total_camps"] > 0)
    ].copy()

    non_mr_activity = non_mr_activity[
        ["empId", "mr_name", "mr_designation", "hq",
         "overall_total_tests", "overall_total_camps"]
    ]

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl",
        mode="a"
    ) as writer:
        non_mr_activity.to_excel(writer, sheet_name="Non MR Tests", index=False)

    print("Non-MR Tests sheet built:", non_mr_activity.shape)

    # 9️⃣ CREATE ONE SHEET PER RBM WITH FULL HIERARCHY (MR → ABM → RBM)
    rbm_list = waterfall_df["rbm_name"].fillna("").unique().tolist()

    used_sheet_names = set()
    blank_count = 0

    with pd.ExcelWriter(output_file, engine="openpyxl", mode="a") as writer:

        for rbm in rbm_list:

            raw_name = rbm if isinstance(rbm, str) else ""

            rbm_block = waterfall_df[
                (waterfall_df["rbm_name"].fillna("") == raw_name) &
                (waterfall_df["designation"].isin(["mr", "abm", "rbm"]))
            ].copy()

            if rbm_block.empty:
                continue

            if raw_name.strip() == "":
                blank_count += 1
                safe_sheet_name = " " * blank_count
            else:
                safe_sheet_name = raw_name.strip()[:31]
                for bad, rep in [("/", "-"), ("\\", "-"), ("[", "("), ("]", ")"),
                                 ("?", ""), ("*", ""), (":", "-")]:
                    safe_sheet_name = safe_sheet_name.replace(bad, rep)

            original = safe_sheet_name
            counter = 1
            while safe_sheet_name in used_sheet_names:
                suffix = f"_{counter}"
                safe_sheet_name = original[:31 - len(suffix)] + suffix
                counter += 1

            used_sheet_names.add(safe_sheet_name)

            rbm_block.to_excel(writer, sheet_name=safe_sheet_name, index=False)

            print(f"Created RBM Sheet: '{safe_sheet_name}'  (RBM: '{raw_name}')")

    # 🔟 APPLY PROFESSIONAL EXCEL STYLING (CUSTOM COLORS)
    wb = load_workbook(output_file)

    COLOR_HEADER = "9DC3E6"   # Blue header
    COLOR_MR     = "FFFFFF"   # White
    COLOR_ABM    = "FFF7CE"   # Light Yellow
    COLOR_RBM    = "F8CBAD"   # Light Red
    COLOR_SM     = "E4DFEC"   # Light Purple

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        header = True

        header_row = [cell.value for cell in ws[1]]
        if "designation" in [str(h).lower() for h in header_row]:
            desig_col = header_row.index("designation") + 1
        else:
            desig_col = None

        for row in ws.iter_rows():
            for cell in row:

                col_letter = cell.column_letter
                cell_len = len(str(cell.value)) if cell.value else 0
                ws.column_dimensions[col_letter].width = max(
                    ws.column_dimensions[col_letter].width or 10,
                    cell_len + 2
                )

                if header:
                    cell.font = Font(bold=True, color="000000")
                    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    continue

                if desig_col:
                    desig_value = ws.cell(row=cell.row, column=desig_col).value
                    d = str(desig_value).strip().lower() if desig_value else ""

                    if d == "mr":
                        fill_color = COLOR_MR
                        bold = False
                    elif d == "abm":
                        fill_color = COLOR_ABM
                        bold = True
                    elif d == "rbm":
                        fill_color = COLOR_RBM
                        bold = True
                    elif d == "sm":
                        fill_color = COLOR_SM
                        bold = True
                    else:
                        fill_color = COLOR_MR
                        bold = False

                    cell.fill = PatternFill("solid", fgColor=fill_color)
                    cell.font = Font(bold=bold)

                cell.border = border
                cell.alignment = Alignment(vertical="center")

            header = False

    wb.save(output_file)
    print(f"🎨 Excel Styling Applied Successfully for {output_file}")


# ------------------------------------------------------------
# 3️⃣ LOOP THROUGH ALL PHARMAS IN THE LIST
# ------------------------------------------------------------
for _, row in pharma_df.iterrows():
    aid = row["aid"]
    pharma_name = row["pharma_name"]
    json_file = row["json_file"]
    prescription_key = row.get("prescription_key", "")

    generate_pharma_report(
        aid=aid,
        pharma_name=pharma_name,
        json_file=json_file,
        prescription_key=prescription_key,
        conn=conn,
        bucket=bucket,
        start_date=START_DATE,
        end_date=END_DATE
    )

print("\n✅✅ All pharma reports generated.")
